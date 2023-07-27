import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side 
from PyQt5.QtWidgets import QDialog 
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import pyqtSignal, QObject,QTimer
from PyQt5 import QtCore, QtWidgets,QtGui
import threading

class SignalHandler(QObject):
    upload_finished = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

class consolidado(QDialog):
    def __init__(self):
        super().__init__()
        self.setModal(True)
        self.window()

    #############FUNCIONES#############
    def window(self):
        self.setObjectName("ADVERTENCIA")
        self.setWindowTitle("ADVERTENCIA")
        self.resize(400, 300)
        self.setStyleSheet("background-color:#cbcbcb;")
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setWindowOpacity(1)    
        self.pushButton = QtWidgets.QPushButton(self)
        self.pushButton.setText("EJECUTAR PROGRAMA")
        self.pushButton.setGeometry(QtCore.QRect(200, 120, 121, 51))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.setStyleSheet("background-color:rgb(187, 17, 17);")
        self.pushButton.clicked.connect(self.ejectutar)
        self.checkBox = QtWidgets.QCheckBox(self)
        self.checkBox.setText("Enero")
        self.checkBox.setGeometry(QtCore.QRect(20, 50, 70, 17))
        self.checkBox.setObjectName("checkBox")
        self.checkBox_2 = QtWidgets.QCheckBox(self)
        self.checkBox_2.setText("Febrero")
        self.checkBox_2.setGeometry(QtCore.QRect(20, 70, 70, 17))
        self.checkBox_2.setObjectName("checkBox_2")
        self.checkBox_3 = QtWidgets.QCheckBox(self)
        self.checkBox_3.setText("Marzo")
        self.checkBox_3.setGeometry(QtCore.QRect(20, 90, 70, 17))
        self.checkBox_3.setObjectName("checkBox_3")
        self.checkBox_4 = QtWidgets.QCheckBox(self)
        self.checkBox_4.setText("Junio")
        self.checkBox_4.setGeometry(QtCore.QRect(20, 150, 70, 17))
        self.checkBox_4.setObjectName("checkBox_4")
        self.checkBox_5 = QtWidgets.QCheckBox(self)
        self.checkBox_5.setText("Mayo")
        self.checkBox_5.setGeometry(QtCore.QRect(20, 130, 70, 17))
        self.checkBox_5.setObjectName("checkBox_5")
        self.checkBox_6 = QtWidgets.QCheckBox(self)
        self.checkBox_6.setText("Abril")
        self.checkBox_6.setGeometry(QtCore.QRect(20, 110, 70, 17))
        self.checkBox_6.setObjectName("checkBox_6")
        self.checkBox_7 = QtWidgets.QCheckBox(self)
        self.checkBox_7.setText("Septiembre")
        self.checkBox_7.setGeometry(QtCore.QRect(20, 210, 81, 17))
        self.checkBox_7.setObjectName("checkBox_7")
        self.checkBox_8 = QtWidgets.QCheckBox(self)
        self.checkBox_8.setText("Agosto")
        self.checkBox_8.setGeometry(QtCore.QRect(20, 190, 70, 17))
        self.checkBox_8.setObjectName("checkBox_8")
        self.checkBox_9 = QtWidgets.QCheckBox(self)
        self.checkBox_9.setText("Julio")
        self.checkBox_9.setGeometry(QtCore.QRect(20, 170, 70, 17))
        self.checkBox_9.setObjectName("checkBox_9")
        self.checkBox_10 = QtWidgets.QCheckBox(self)
        self.checkBox_10.setText("Diciembre")
        self.checkBox_10.setGeometry(QtCore.QRect(20, 270, 70, 17))
        self.checkBox_10.setObjectName("checkBox_10")
        self.checkBox_11 = QtWidgets.QCheckBox(self)
        self.checkBox_11.setText("Noviembre")
        self.checkBox_11.setGeometry(QtCore.QRect(20, 250, 70, 17))
        self.checkBox_11.setObjectName("checkBox_11")
        self.checkBox_12 = QtWidgets.QCheckBox(self)
        self.checkBox_12.setText("Octubre")
        self.checkBox_12.setGeometry(QtCore.QRect(20, 230, 70, 17))
        self.checkBox_12.setObjectName("checkBox_12")
        self.label = QtWidgets.QLabel(self)
        self.label.setGeometry(QtCore.QRect(90, 10, 221, 16))
        font = QtGui.QFont()
        font.setPointSize(15)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label.setText("Erosión de materiales")

        QtCore.QMetaObject.connectSlotsByName(self)
        self.signal_handler = SignalHandler()
        self.signal_handler.upload_finished.connect(self.finish)


        self.checkBox.stateChanged.connect(self.Eleccion)
        self.checkBox_2.stateChanged.connect(self.Eleccion)
        self.checkBox_3.stateChanged.connect(self.Eleccion)
        self.checkBox_4.stateChanged.connect(self.Eleccion)
        self.checkBox_6.stateChanged.connect(self.Eleccion)
        self.checkBox_7.stateChanged.connect(self.Eleccion)
        self.checkBox_8.stateChanged.connect(self.Eleccion)
        self.checkBox_9.stateChanged.connect(self.Eleccion)
        self.checkBox_10.stateChanged.connect(self.Eleccion)
        self.checkBox_11.stateChanged.connect(self.Eleccion)
        self.checkBox_12.stateChanged.connect(self.Eleccion)

    def show_finish(self):
        QTimer.singleShot(0, self.finish) 
    def finish(self):
        QMessageBox.information(self, "Proceso finalizado", "Se ha procesado el archivo correctamente.")
   
    def ejectutar(self):
        try:
            upload_Thread=threading.Thread(target=self.formarto())
            upload_Thread.start()
            upload_Thread.join()
            self.signal_handler.upload_finished.emit()
            self.close()
        except PermissionError as P:
            print(f"Error: {P}")
            QMessageBox.warning(self, "Error", "Por favor cierre el archivo.")
    def formatear_dinero(self,valor):
        return "${:,.2f}".format(valor)
    def Eleccion(self):
        checkboxes = [self.checkBox, self.checkBox_2, self.checkBox_3, self.checkBox_4, self.checkBox_5, self.checkBox_6, self.checkBox_7, self.checkBox_8, self.checkBox_9, self.checkBox_10, self.checkBox_11, self.checkBox_12]
        checkboxes_seleccionados = [checkbox for checkbox in checkboxes if checkbox.isChecked()]
        check=True
        if len(checkboxes_seleccionados) >= 2:
            print("Más de dos checkboxes seleccionados:")
            for checkbox in checkboxes_seleccionados:
                print(checkbox.text())
                check=False 
            return check

        else:
            for checkbox in checkboxes:
                if checkbox.isChecked():
                    print(f"Checkbox '{checkbox.text()}' seleccionado")
                    check=checkbox.text()
            return check

    def sumar_o_conservar(self,row):
        if row['CODIGO SAP'] == row['CÓDIGO SAP']:
            return row['TOTAL 3 CAPAS']
        else:
            return row['TOTAL 3 CAPAS'] + row['TOTAL 3 CAPASS']
    def formarto(self):
        name_file=""
        #Rutas de archivos
        path_consolidado_compras='C://Users/IC0167A/Desktop/Materiales_consolidado/05_03 Consolidado Compras en Tránsito (1).xlsx'
        path_MM='C://Users/IC0167A/Desktop/Materiales_consolidado/05. MM_Infr.Fija_Movil_May2023.xlsb'
        path_inventario='C://Users/IC0167A/Desktop/Materiales_consolidado/INVENTARIO  A MAYO 2023 (FOTO).xlsb'
        path_modelo_materiales='C://Users/IC0167A/Desktop/Materiales_consolidado/Modelo de Materiales de redes HFC  MODELO 6.0  BG.xlsb'
        path_familia='C://Users/IC0167A/Desktop/Materiales_consolidado/Familia.xlsx'
        #Lectura de archivos
        db_MM=pd.read_excel(path_MM,sheet_name='MM  INF FIJ_MOV',skiprows=6)
        db_inventario=pd.read_excel(path_inventario,sheet_name='TD BASE GEN',skiprows=2)
        db_con_compras=pd.read_excel(path_consolidado_compras,sheet_name='Hoja1',skiprows=2)
        db_model_materiales=pd.read_excel(path_modelo_materiales,sheet_name='BASE GENERAL',skiprows=3)
        db_familia=pd.read_excel(path_familia,skiprows=4)
        #fitrado de columnas
        values_model_materiales = db_model_materiales.drop(['Unnamed: 0','Segmento','Suma de TOTAL'],axis=1)
        values_inventario=db_inventario.loc[:,['Etiquetas de fila','TOTAL 3 CAPAS']]
        values_conso_compras=db_con_compras.loc[:,['Etiquetas de fila','Fecha Estimada Disponibilidad en Aliado','Suma de Cantidad por entregar']]
        values_MM=db_MM.loc[:,['CODIGO SAP','PRECIO UNITARIO']]
        values_familia=db_familia.loc[:,['Código SAP','Descripción','FAMILIA','CLASIFICACION','CODIGO DEFINITIVO','Código SAP 2']]
        #Renombrar columnas
        values_model_materiales=values_model_materiales.rename(columns={"CODIGO":"CÓDIGO SAP"})
        values_model_materiales['validacion']='ESTO SI'
        values_inventario2=values_inventario.rename(columns={"Etiquetas de fila":"CODIGO SAP"})
        values_inventario=values_inventario.rename(columns={"Etiquetas de fila":"CÓDIGO SAP"})      
        values_conso_compras=values_conso_compras.rename(columns={"Etiquetas de fila":"CODIGO SAP"})
        values_familia=values_familia.rename(columns={'CODIGO DEFINITIVO':'CODIGO SAP'})
        values_familia=values_familia.rename(columns={'Código SAP':'CÓDIGO SAP'})
        v_proyecto=values_model_materiales.loc[:,['PROYECTO','CÓDIGO SAP']]
        print(f"fecha==>{values_conso_compras}")
        ##FECHA
        values_conso_compras=values_conso_compras.sort_values(['Fecha Estimada Disponibilidad en Aliado'])
        print(len(values_conso_compras))
        #values_conso_compras.drop_duplicates(['CODIGO SAP'],inplace=True,keep='last')
        print(len(values_conso_compras))
        print(values_conso_compras)
        # Agrupar por "CÓDIGO SAP" y sumar los meses
        df_agrupado = values_model_materiales.groupby("CÓDIGO SAP")[["ENE 2023", "FEB 2023", "MAR 2023", "ABR 2023", "MAY 2023", "JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]].sum()
        # Combinar el DataFrame agrupado con el DataFrame original
        df_agrupado2 = df_agrupado.reset_index()
        #Se une la columna de PROYECTO
        df_agrupado_model_materiales=df_agrupado2.merge(v_proyecto,on='CÓDIGO SAP',how='left')
        df_agrupado_model_materiales=df_agrupado_model_materiales.drop_duplicates('CÓDIGO SAP')
        #!Suma de inventario
        inv1=df_agrupado_model_materiales.merge(values_familia,on='CÓDIGO SAP',how='left').merge(values_inventario,on='CÓDIGO SAP',how='left')
        inv1=inv1.loc[:,['CÓDIGO SAP','TOTAL 3 CAPAS']]
        inv1=inv1.rename(columns={'TOTAL 3 CAPAS':'TOTAL 3 CAPASS'})
        inv2=df_agrupado_model_materiales.merge(values_familia,on='CÓDIGO SAP',how='left').merge(values_inventario2,on='CODIGO SAP',how='left')
        inv2=inv2.loc[:,['CODIGO SAP','TOTAL 3 CAPAS']]
        #MERGE de todos los archivos
        con=df_agrupado_model_materiales.merge(values_familia,on='CÓDIGO SAP',how='left').merge(inv1,on='CÓDIGO SAP',how='left').merge(inv2,on='CODIGO SAP',how='left').merge(values_conso_compras,on='CODIGO SAP',how='left').merge(values_MM,on='CODIGO SAP',how='left')
        con=con.drop_duplicates()
        con['TOTAL 3 CAPAS']=con.apply(self.sumar_o_conservar,axis=1)
        con.drop(['TOTAL 3 CAPASS'], axis=1, inplace=True)
        #REORDENAR COLUMNAS
        con=con.loc[:,['CÓDIGO SAP','CODIGO SAP','PROYECTO','Descripción','CLASIFICACION','FAMILIA',"ENE 2023", "FEB 2023", "MAR 2023", "ABR 2023", "MAY 2023", "JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023",'TOTAL 3 CAPAS','Suma de Cantidad por entregar','PRECIO UNITARIO','Fecha Estimada Disponibilidad en Aliado']]
        
        mes=self.Eleccion()
        print(mes)
        ##################################################################ENERO##################################################################
        if(mes=="Enero"):
            suma_mes=["ENE 2023", "FEB 2023", "MAR 2023", "ABR 2023", "MAY 2023","JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL ENE-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total ENE_INV']=con['TOTAL 3 CAPAS']-con['ENE 2023'].round()
            con['Total FEB_INV']=con['Total ENE_INV']-con['FEB 2023'].round()
            con['Total MAR_INV']=con['Total FEB_INV']-con['MAR 2023'].round()
            con['Total ABR_INV']=con['Total MAR_INV']-con['ABR 2023'].round()
            con['Total MAY_INV']=con['Total ABR_INV']-con['MAY 2023'].round()
            con['Total JUN_INV']=con['Total MAY_INV']-con['JUN 2023'].round()
            con['Total JUL_INV']=con['Total JUN_INV']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total ENE_ERO']=con['PRECIO UNITARIO']*con['Total ENE_INV']
            con['Total FEB_ERO']=con['PRECIO UNITARIO']*con['Total FEB_INV']
            con['Total MAR_ERO']=con['PRECIO UNITARIO']*con['Total MAR_INV']
            con['Total ABR_ERO']=con['PRECIO UNITARIO']*con['Total ABR_INV']
            con['Total MAY_ERO']=con['PRECIO UNITARIO']*con['Total MAY_INV']
            con['Total JUN_ERO']=con['PRECIO UNITARIO']*con['Total JUN_INV']
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Enero.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AI','AW']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL ENE-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras
        ##################################################################FEBRERO##################################################################    
        elif(mes=="Febrero"):
            suma_mes=["FEB 2023", "MAR 2023", "ABR 2023", "MAY 2023","JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL FEB-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total FEB_INV']=con['TOTAL 3 CAPAS']-con['FEB 2023'].round()
            con['Total MAR_INV']=con['Total FEB_INV']-con['MAR 2023'].round()
            con['Total ABR_INV']=con['Total MAR_INV']-con['ABR 2023'].round()
            con['Total MAY_INV']=con['Total ABR_INV']-con['MAY 2023'].round()
            con['Total JUN_INV']=con['Total MAY_INV']-con['JUN 2023'].round()
            con['Total JUL_INV']=con['Total JUN_INV']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total FEB_ERO']=con['PRECIO UNITARIO']*con['Total FEB_INV']
            con['Total MAR_ERO']=con['PRECIO UNITARIO']*con['Total MAR_INV']
            con['Total ABR_ERO']=con['PRECIO UNITARIO']*con['Total ABR_INV']
            con['Total MAY_ERO']=con['PRECIO UNITARIO']*con['Total MAY_INV']
            con['Total JUN_ERO']=con['PRECIO UNITARIO']*con['Total JUN_INV']
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Febrero.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AI','AU']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL FEB-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras
        ##################################################################MARZO###################################################################
        elif(mes=="Marzo"):
            suma_mes=["MAR 2023", "ABR 2023", "MAY 2023","JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL MAR-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total MAR_INV']=con['TOTAL 3 CAPAS']-con['MAR 2023'].round()
            con['Total ABR_INV']=con['Total MAR_INV']-con['ABR 2023'].round()
            con['Total MAY_INV']=con['Total ABR_INV']-con['MAY 2023'].round()
            con['Total JUN_INV']=con['Total MAY_INV']-con['JUN 2023'].round()
            con['Total JUL_INV']=con['Total JUN_INV']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total MAR_ERO']=con['PRECIO UNITARIO']*con['Total MAR_INV']
            con['Total ABR_ERO']=con['PRECIO UNITARIO']*con['Total ABR_INV']
            con['Total MAY_ERO']=con['PRECIO UNITARIO']*con['Total MAY_INV']
            con['Total JUN_ERO']=con['PRECIO UNITARIO']*con['Total JUN_INV']
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)     
            name_file="Erosión Marzo.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AH','AS']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL MAR-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras      
        ##################################################################ABRIL###################################################################
        elif(mes=="Abril"):
            suma_mes=["ABR 2023", "MAY 2023","JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL ABR-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total ABR_INV']=con['TOTAL 3 CAPAS']-con['ABR 2023'].round()
            con['Total MAY_INV']=con['Total ABR_INV']-con['MAY 2023'].round()
            con['Total JUN_INV']=con['Total MAY_INV']-con['JUN 2023'].round()
            con['Total JUL_INV']=con['Total JUN_INV']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total ABR_ERO']=con['PRECIO UNITARIO']*con['Total ABR_INV']
            con['Total MAY_ERO']=con['PRECIO UNITARIO']*con['Total MAY_INV']
            con['Total JUN_ERO']=con['PRECIO UNITARIO']*con['Total JUN_INV']
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)         
            name_file="Erosión Abril.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AG','AQ']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL ABR-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras 
        ##################################################################MAYO####################################################################
        elif(mes=="Mayo"):
            suma_mes=["MAY 2023","JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL MAY-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total MAY_INV']=con['TOTAL 3 CAPAS']-con['MAY 2023'].round()
            con['Total JUN_INV']=con['Total MAY_INV']-con['JUN 2023'].round()
            con['Total JUL_INV']=con['Total JUN_INV']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total MAY_ERO']=con['PRECIO UNITARIO']*con['Total MAY_INV']
            con['Total JUN_ERO']=con['PRECIO UNITARIO']*con['Total JUN_INV']
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Mayo.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AF','AO']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL ABR-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras     
        ###################################################################JUNIO##################################################################
        elif(mes=="Junio"):
            suma_mes=["JUN 2023", "JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL JUN-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total JUN_INV']=con['TOTAL 3 CAPAS']-con['JUN 2023'].round()
            con['Total JUL_INV']=con['Total JUN_INV']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total JUN_ERO']=con['PRECIO UNITARIO']*con['Total JUN_INV']
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero) 
            name_file="Erosión Junio.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AE','AM']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL JUN-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras    
        ##################################################################JULIO###################################################################
        elif(mes=="Julio"):
            suma_mes=["JUL 2023", "AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL JUL-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total JUL_INV']=con['TOTAL 3 CAPAS']-con['JUL 2023'].round()
            con['Total AGO_INV']=con['Total JUL_INV']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total JUL_ERO']=con['PRECIO UNITARIO']*con['Total JUL_INV']
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero) 
            name_file="Erosión Julio.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AD','AK']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL JUL-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras 
        ##################################################################AGOSTO##################################################################
        elif(mes=="Agosto"):
            suma_mes=["AGO 2023", "SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL AGO-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total AGO_INV']=con['TOTAL 3 CAPAS']-con['AGO 2023'].round()
            con['Total SEP_INV']=con['Total AGO_INV']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total AGO_ERO']=con['PRECIO UNITARIO']*con['Total AGO_INV']
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Agosto.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AC','AI']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL AGO-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras 
        ##################################################################SEPTIEMBRE##############################################################
        elif(mes=="Septiembre"):
            suma_mes=["SEP 2023", "OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL SEP-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total SEP_INV']=con['TOTAL 3 CAPAS']-con['SEP 2023'].round()
            con['Total OCT_INV']=con['Total SEP_INV']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total SEP_ERO']=con['PRECIO UNITARIO']*con['Total SEP_INV']
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Septiembre.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AB','AG']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL SEP-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras 
        ##################################################################OCTUBRE#################################################################
        elif(mes=="Octubre"):
            suma_mes=["OCT 2023", "NOV 2023", "DIC 2023"]
            con['TOTAL OCT-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total OCT_INV']=con['TOTAL 3 CAPAS']-con['OCT 2023'].round()
            con['Total NOV_INV']=con['Total OCT_INV']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total OCT_ERO']=con['PRECIO UNITARIO']*con['Total OCT_INV']
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Octubre.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','AA','AE']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL OCT-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras 
        ##################################################################NOVIEMBRE###############################################################
        elif(mes=="Noviembre"):
            suma_mes=["NOV 2023", "DIC 2023"]
            con['TOTAL NOV-DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total NOV_INV']=con['TOTAL 3 CAPAS']-con['NOV 2023'].round()
            con['Total DIC_INV']=con['Total NOV_INV']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total NOV_ERO']=con['PRECIO UNITARIO']*con['Total NOV_INV']
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero)
            name_file="Erosión Noviembre.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','Z','AC']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL NOV-DIC-23','Disponible o faltante','EROSION Total X Proyecto']  # Lista de cabeceras 
        ##################################################################DICIEMBRE###############################################################
        elif(mes=="Diciembre"):  
            suma_mes=["DIC 2023"]
            con['TOTAL DIC-23']=con[suma_mes].sum(axis=1,skipna=True).round()
            #########################VALORES EN NULL#########################
            con['TOTAL 3 CAPAS']=con['TOTAL 3 CAPAS'].fillna(value=0)
            con['PRECIO UNITARIO']=con['PRECIO UNITARIO'].fillna(value=0)
            ##########################INVENTARIO#############################
            con['Total DIC_INV']=con['TOTAL 3 CAPAS']-con['DIC 2023'].round()
            con['Total x cantidad']=con['Total DIC_INV']
            ###########################EROSION###############################
            con['Total DIC_ERO']=con['PRECIO UNITARIO']*con['Total DIC_INV']
            con['EROSION Total X Proyecto']=con['Total DIC_ERO']
            con['EROSION Total X Proyecto'] = con['EROSION Total X Proyecto'].map(self.formatear_dinero) 
            name_file="Erosión Diciembre.xlsx"
            columnas_origen = ['A', 'B', 'C','D','S','W','Y','AA']  # Lista de las letras de las columnas de origen 
            columnas_destino = ['A', 'B', 'C','D','E','F','G','H']  # Lista de las letras de las columnas de destino 
            cabeceras = ['CÓDIGO SAP', 'CODIGO SAP', 'PROYECTO','Descripción','Inventario','TOTAL DIC-23','Total x cantidad','EROSION Total X Proyecto']  # Lista de cabeceras 
        else:
           QMessageBox.warning(self, "Error", "Por favor, solo seleccione un mes.") 
        ###Operaciones
        ######################################################SUMA MES A MES#############################################################

        with pd.ExcelWriter(name_file,engine='xlsxwriter') as excelfile:
            con.to_excel(excelfile,sheet_name='Consolidado',index=False)
            
        # Cargar el archivo Excel
        workbook = load_workbook(name_file)

        # Obtener la hoja
        worksheet = workbook['Consolidado']
        #Nueva hoja
        hoja_destino = workbook.create_sheet(title='Resumen')

        for columna_origen, columna_destino in zip(columnas_origen, columnas_destino):
            datos_origen = worksheet[columna_origen]
            for fila, celda in enumerate(datos_origen, start=1):
                celda_destino = hoja_destino.cell(row=fila, column=ord(columna_destino) - 64)
                celda_destino.value = celda.value

        for indice, cabecera in enumerate(cabeceras, start=1):
            celda_cabecera = hoja_destino.cell(row=1, column=indice)
            celda_cabecera.value = cabecera 

            # Establecer formato en negrita
            font = Font(bold=True) 
            celda_cabecera.font = font

            # Establecer bordes
            border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"),
                            left=Side(border_style="thin"), right=Side(border_style="thin"))
            celda_cabecera.border = border


        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1  # Ajusta el ancho multiplicando la longitud por un factor
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        for column in hoja_destino.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1  # Ajusta el ancho multiplicando la longitud por un factor
            hoja_destino.column_dimensions[column[0].column_letter].width = adjusted_width

        # Guardar los cambios en el archivo Excel
        workbook.save(name_file)
        print(df_agrupado_model_materiales)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    ui = consolidado()
    ui.show()
    sys.exit(app.exec_())
    